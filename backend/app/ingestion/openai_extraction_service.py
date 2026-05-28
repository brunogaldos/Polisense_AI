"""OpenAI-powered document extraction — Python port of openaiExtractionService.ts.

Supported input types:
  • PDF  – extracts embedded text via pypdf; for scanned / malformed PDFs,
           uploads the file to OpenAI's Files API and reads it with GPT-4o via
           the Responses API (no local rendering needed).
  • Images – PNG, JPG, JPEG, WEBP, GIF, TIFF – sent directly to GPT-4o vision.
  • Spreadsheets – xlsx-family parsed locally via openpyxl into row-bounded
           Markdown tables (see extract_from_spreadsheet for rationale).

All methods are synchronous (the OpenAI SDK + pypdf + openpyxl calls are
blocking) — call them via asyncio.to_thread from async handlers.
"""

import base64
import io
import logging
import os
import re
from datetime import date, datetime
from typing import Optional

logger = logging.getLogger(__name__)

# ─── Configuration ──────────────────────────────────────────────────────────

EXTRACTION_MODEL = os.getenv("EXTRACTION_MODEL", "gpt-4o")

# Minimum average characters per page for a PDF to be treated as text-based.
# Below this threshold we fall back to the vision (OCR) path.
MIN_CHARS_PER_PAGE = 80

# Max rows we'll render per sheet before truncating with a marker.
SPREADSHEET_MAX_ROWS_PER_SHEET = int(os.getenv("SPREADSHEET_MAX_ROWS_PER_SHEET") or 5000)

# Hard cap on total markdown bytes produced for a single workbook.
SPREADSHEET_MAX_MARKDOWN_BYTES = int(os.getenv("SPREADSHEET_MAX_MARKDOWN_BYTES") or 8 * 1024 * 1024)

# Sheet names that almost always carry non-content (lookup tables, build flags,
# internal notes). Skipped by default to keep retrieval focused.
_JUNK_SHEET_NAME = re.compile(
    r"^(config|configuration|settings|metadata|hidden|sheet0|_.*|.*\$.*|"
    r"drop[-_ ]?downs?|lookups?|reference|references)$",
    re.IGNORECASE,
)


def _detect_tables(md: str) -> bool:
    return bool(re.search(r"\|.+\|.+\|", md))


def _detect_formulas(md: str) -> bool:
    return bool(re.search(r"\$[^$]+\$|\$\$[^$]+\$\$|\\begin\{|\\end\{", md))


class OpenAIExtractionService:
    _client = None

    @classmethod
    def _get_client(cls):
        if cls._client is None:
            from openai import OpenAI

            api_key = os.getenv("OPENAI_API_KEY") or os.getenv("AI_MODEL_API_KEY")
            if not api_key:
                raise RuntimeError("OPENAI_API_KEY is not configured")
            cls._client = OpenAI(api_key=api_key)
        return cls._client

    # ─── Public entry point ──────────────────────────────────────────────────

    @classmethod
    def extract_from_file(cls, file_buffer: bytes, file_name: str, mime_type: str) -> dict:
        """Extract content from a PDF or image buffer as Markdown."""
        lower = file_name.lower()
        is_pdf = mime_type == "application/pdf" or lower.endswith(".pdf")
        is_image = bool(
            re.match(r"^image/(png|jpeg|gif|webp|tiff)", mime_type or "", re.IGNORECASE)
        ) or bool(re.search(r"\.(png|jpe?g|gif|webp|tiff?)$", file_name, re.IGNORECASE))

        if is_pdf:
            return cls._extract_from_pdf(file_buffer, file_name)
        if is_image:
            return cls.extract_from_image(file_buffer, file_name, mime_type)
        raise ValueError(
            f"Unsupported file type: {mime_type}. Supported: PDF, PNG, JPG, JPEG, GIF, WEBP, TIFF"
        )

    # ─── PDF extraction ────────────────────────────────────────────────────────

    @classmethod
    def _extract_from_pdf(cls, pdf_buffer: bytes, file_name: str) -> dict:
        logger.info("Extracting PDF: %s (%.2f MB)", file_name, len(pdf_buffer) / 1024 / 1024)

        # 1. Try embedded-text extraction (fast, free).
        raw_text = ""
        pages = 1
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(pdf_buffer))
            pages = len(reader.pages) or 1
            raw_text = "\n".join((page.extract_text() or "") for page in reader.pages)
            logger.info("pypdf: %d pages, %d chars total", pages, len(raw_text))
        except Exception as err:  # noqa: BLE001
            logger.warning("pypdf failed: %s. Falling back to vision path.", err)

        avg_chars_per_page = len(raw_text) / max(1, pages)
        logger.info(
            "Avg chars/page: %.0f (threshold: %d)", avg_chars_per_page, MIN_CHARS_PER_PAGE
        )

        # 2. Route based on text density.
        if avg_chars_per_page >= MIN_CHARS_PER_PAGE:
            logger.info("Text-rich PDF → returning raw text directly")
            return {
                "text": raw_text,
                "markdown": raw_text,
                "pages": pages,
                "extractionMethod": "OPENAI",
                "hasTables": _detect_tables(raw_text),
                "hasFormulas": _detect_formulas(raw_text),
            }

        logger.info("Sparse text — using GPT-4o vision (OCR) path")
        return cls._extract_pdf_with_vision(pdf_buffer, file_name, pages, raw_text)

    @classmethod
    def _extract_pdf_with_vision(
        cls, pdf_buffer: bytes, file_name: str, estimated_pages: int, fallback_text: str
    ) -> dict:
        """Upload the PDF to OpenAI's Files API and let GPT-4o read it natively
        via the Responses API. The uploaded PDF is kept alive on success so the
        caller can attach it via `input_file`; only deleted on a hard failure."""
        client = cls._get_client()
        uploaded_file_id: Optional[str] = None
        succeeded = False
        try:
            logger.info("Uploading PDF to OpenAI Files API for vision extraction…")
            uploaded = client.files.create(
                file=(file_name, pdf_buffer, "application/pdf"), purpose="user_data"
            )
            uploaded_file_id = uploaded.id
            logger.info("PDF uploaded as %s — calling GPT-4o for extraction", uploaded.id)

            resp = client.responses.create(
                model=EXTRACTION_MODEL,
                input=[
                    {
                        "role": "user",
                        "content": [
                            # `detail` mirrors the Node reference implementation.
                            {"type": "input_file", "file_id": uploaded.id, "detail": "high"},
                            {
                                "type": "input_text",
                                "text": (
                                    "Extract ALL text from this document as clean Markdown. "
                                    "Preserve tables (Markdown pipe syntax), lists, headings, "
                                    "and numeric data. Return ONLY the extracted Markdown — no "
                                    "commentary, no preamble."
                                ),
                            },
                        ],
                    }
                ],
            )

            markdown = getattr(resp, "output_text", "") or ""
            if not markdown or len(markdown) < 10:
                raise RuntimeError("GPT-4o returned empty extraction")

            logger.info(
                "Vision extraction done: %d chars from ~%d page(s)", len(markdown), estimated_pages
            )
            succeeded = True
            return {
                "text": markdown,
                "markdown": markdown,
                "pages": estimated_pages,
                "extractionMethod": "OPENAI",
                "hasTables": _detect_tables(markdown),
                "hasFormulas": _detect_formulas(markdown),
                "visionUploadedFileId": uploaded_file_id,
            }
        except Exception as err:  # noqa: BLE001
            logger.error("OpenAI vision extraction failed: %s", err)
            if len(fallback_text) > 10:
                logger.info("Using raw pypdf fallback (%d chars)", len(fallback_text))
                succeeded = True  # keep the file: still attachable for chat-time visual reads
                return {
                    "text": fallback_text,
                    "markdown": fallback_text,
                    "pages": estimated_pages,
                    "extractionMethod": "OPENAI",
                    "hasTables": False,
                    "hasFormulas": False,
                    "visionUploadedFileId": uploaded_file_id,
                }
            raise RuntimeError(f"Could not extract PDF text and no fallback available: {err}")
        finally:
            # Only delete on hard failure with no fallback — otherwise the upload
            # is intentionally kept so the chatbot can attach it as `input_file`.
            if uploaded_file_id and not succeeded:
                try:
                    client.files.delete(uploaded_file_id)
                except Exception as e:  # noqa: BLE001
                    logger.warning("Could not delete temp file %s: %s", uploaded_file_id, e)

    @classmethod
    def upload_for_attachment(cls, file_buffer: bytes, file_name: str, mime_type: str) -> str:
        """Upload a file's original bytes to OpenAI Files API with
        purpose='user_data' so it can be attached to a Responses API call via
        `input_file`. The caller owns the file's lifecycle."""
        client = cls._get_client()
        uploaded = client.files.create(file=(file_name, file_buffer, mime_type), purpose="user_data")
        logger.info("Uploaded %r for attachment: %s", file_name, uploaded.id)
        return uploaded.id

    # ─── Spreadsheet extraction (xlsx / xlsm / xltm / xltx) ────────────────────

    @classmethod
    def extract_from_spreadsheet(cls, buffer: bytes, file_name: str) -> dict:
        """Parse an OOXML workbook locally to row-bounded Markdown tables.

        Why local extraction instead of dumping the file to the vector store:
        OpenAI's default chunker is text-oriented and can split a row mid-cell
        (losing column→value pairing), indexes noisy helper/hidden sheets, and
        silently truncates very large spreadsheets. Going via Markdown gives
        deterministic sheet boundaries and visibility into what's indexed.
        """
        logger.info(
            "openpyxl: Extracting spreadsheet: %s (%.2f MB)", file_name, len(buffer) / 1024 / 1024
        )
        try:
            from openpyxl import load_workbook

            # data_only=True → use cached computed values for formula cells
            # (mirrors the Node "prefer computed result" behaviour).
            workbook = load_workbook(
                io.BytesIO(buffer), read_only=True, data_only=True, keep_links=False
            )
        except Exception as err:  # noqa: BLE001
            raise ValueError(
                f'Could not parse "{file_name}" as Office Open XML (xlsx-family). '
                f"If this is a legacy .xls or binary .xlsb file, please save it as .xlsx "
                f"and re-upload. ({err})"
            )

        sections: list[str] = []
        total_bytes = 0
        total_rows = 0
        skipped_sheets = 0
        truncated_sheets = 0

        for sheet in workbook.worksheets:
            # openpyxl sheet_state is 'visible' | 'hidden' | 'veryHidden'.
            if sheet.sheet_state and sheet.sheet_state != "visible":
                skipped_sheets += 1
                continue
            if _JUNK_SHEET_NAME.match(sheet.title or ""):
                skipped_sheets += 1
                continue

            markdown, rows_rendered, truncated = cls._render_sheet(sheet)
            if not markdown:
                skipped_sheets += 1
                continue
            section = f"# Sheet: {sheet.title}\n\n{markdown}\n"
            if total_bytes + len(section) > SPREADSHEET_MAX_MARKDOWN_BYTES:
                sections.append(
                    f"\n_⚠️ Remaining sheets not indexed: workbook exceeds the "
                    f"{SPREADSHEET_MAX_MARKDOWN_BYTES // 1024 // 1024} MB markdown budget. "
                    f"Drop unused sheets and re-upload for full coverage._\n"
                )
                break
            sections.append(section)
            total_bytes += len(section)
            total_rows += rows_rendered
            if truncated:
                truncated_sheets += 1

        workbook.close()

        if not sections:
            raise ValueError(
                f'No usable sheets found in "{file_name}" (workbook may be empty, '
                f"password-protected, or contain only hidden/helper sheets)."
            )

        sheet_count = len([s for s in sections if s.startswith("# Sheet:")])
        header_note = (
            f"> Source workbook: `{file_name}` — "
            f"{sheet_count} sheet{'' if sheet_count == 1 else 's'} indexed"
            + (f", {skipped_sheets} skipped (hidden/empty/internal)" if skipped_sheets else "")
            + (
                f", {truncated_sheets} truncated to {SPREADSHEET_MAX_ROWS_PER_SHEET} rows"
                if truncated_sheets
                else ""
            )
            + "."
        )
        markdown = f"# Workbook: {file_name}\n\n{header_note}\n\n" + "\n".join(sections)

        logger.info(
            "Spreadsheet extracted: %d sheets, %d rows, %.1f KB markdown",
            sheet_count,
            total_rows,
            len(markdown) / 1024,
        )

        return {
            "text": markdown,
            "markdown": markdown,
            "pages": sheet_count,
            "extractionMethod": "OPENAI",
            "hasTables": True,
            "hasFormulas": False,
        }

    @classmethod
    def _render_sheet(cls, sheet) -> tuple[str, int, bool]:
        """Render a single worksheet as a Markdown table.
        - Treats the FIRST non-empty row as header.
        - Truncates at SPREADSHEET_MAX_ROWS_PER_SHEET rows with a marker.
        - Escapes pipes so the markdown table doesn't break."""
        rows: list[list[str]] = []
        max_cols = 0
        stopped = False

        for raw in sheet.iter_rows(values_only=True):
            if len(rows) >= SPREADSHEET_MAX_ROWS_PER_SHEET:
                stopped = True
                break
            cells = [cls._cell_to_text(v) for v in raw]
            # Trim trailing empties that exist purely because another row was wider.
            while cells and (cells[-1] == "" or cells[-1] is None):
                cells.pop()
            if not cells:
                continue  # entirely-blank row after trim
            max_cols = max(max_cols, len(cells))
            rows.append(cells)

        if not rows:
            return "", 0, False

        # Normalise width so every row has the same column count.
        for r in rows:
            while len(r) < max_cols:
                r.append("")

        header = rows[0]
        body = rows[1:]
        header_line = "| " + " | ".join(cls._escape_cell(c) for c in header) + " |"
        sep_line = "| " + " | ".join("---" for _ in header) + " |"
        body_lines = ["| " + " | ".join(cls._escape_cell(c) for c in r) + " |" for r in body]
        md = "\n".join([header_line, sep_line, *body_lines])

        trailer = (
            f"\n\n_…truncated at {SPREADSHEET_MAX_ROWS_PER_SHEET} rows; sheet has more data._"
            if stopped
            else ""
        )
        return md + trailer, len(rows), stopped

    @staticmethod
    def _cell_to_text(v) -> str:
        """Convert an openpyxl cell value to a stable text representation.
        With data_only=True + values_only=True, cells arrive as native Python
        types (str/int/float/bool/datetime/None)."""
        if v is None:
            return ""
        if isinstance(v, str):
            return v
        if isinstance(v, bool):
            return "true" if v else "false"
        if isinstance(v, (int, float)):
            return str(v)
        if isinstance(v, (datetime, date)):
            return v.isoformat()
        return str(v)

    @staticmethod
    def _escape_cell(s: str) -> str:
        """Escape markdown-table-breaking characters inside a cell."""
        return re.sub(r"\r?\n", " ", s or "").replace("|", "\\|").strip()

    # ─── Image extraction ──────────────────────────────────────────────────────

    @classmethod
    def extract_from_image(cls, image_buffer: bytes, file_name: str, mime_type: str) -> dict:
        logger.info("Extracting image: %s (%.1f KB)", file_name, len(image_buffer) / 1024)
        client = cls._get_client()
        b64 = base64.b64encode(image_buffer).decode("ascii")
        media_type = mime_type if (mime_type or "").startswith("image/") else "image/png"

        resp = client.chat.completions.create(
            model=EXTRACTION_MODEL,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a document analysis assistant. Analyse the provided image and "
                        "extract all visible text and information. Convert to clean, "
                        "well-structured Markdown preserving tables, lists, headings, and data "
                        "values. Return ONLY the Markdown content — no explanations."
                    ),
                },
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "Extract all content from this image as clean Markdown:"},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{media_type};base64,{b64}",
                                "detail": "high",
                            },
                        },
                    ],
                },
            ],
            temperature=0,
            max_tokens=4096,
        )

        markdown = (resp.choices[0].message.content or "") if resp.choices else ""
        logger.info("Image extraction done: %d chars", len(markdown))
        return {
            "text": markdown,
            "markdown": markdown,
            "pages": 1,
            "extractionMethod": "OPENAI",
            "hasTables": _detect_tables(markdown),
            "hasFormulas": _detect_formulas(markdown),
        }
